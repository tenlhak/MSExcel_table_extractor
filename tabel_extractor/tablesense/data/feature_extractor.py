import pandas as pd
import openpyxl
import numpy as np
import datetime
from typing import Optional



def extract_table_features(excel_path: str, sheet_name: str) -> np.ndarray:
    """
    Extracts feature vectors from an Excel sheet (.xlsx) with robust error handling and sheet name matching.
    
    This function implements several validation strategies to handle common issues in Excel files:
    - Sheet name variations (case sensitivity, whitespace)
    - Various cell formats and properties
    - Error cases and missing attributes
    
    Args:
        excel_path: Path to the Excel file (.xlsx)
        sheet_name: Name of the sheet to process
    
    Returns:
        np.ndarray: A tensor of shape (height, width, 20) containing cell features
    """
    def find_matching_sheet(workbook, target_sheet_name: str) -> str:
        """
        Finds the best matching sheet name in the workbook using multiple matching strategies.
        This helps handle common issues like trailing spaces or case mismatches.
        """
        # First try: exact match
        if target_sheet_name in workbook.sheetnames:
            return target_sheet_name
            
        # Second try: remove whitespace
        stripped_name = target_sheet_name.strip()
        if stripped_name in workbook.sheetnames:
            print(f"Found sheet '{stripped_name}' after stripping whitespace")
            return stripped_name
            
        # Third try: case-insensitive match
        for sheet in workbook.sheetnames:
            if sheet.lower() == target_sheet_name.lower():
                print(f"Found sheet '{sheet}' through case-insensitive match")
                return sheet
                
        # Final try: case-insensitive and whitespace-stripped match
        for sheet in workbook.sheetnames:
            if sheet.lower().strip() == target_sheet_name.lower().strip():
                print(f"Found sheet '{sheet}' through case-insensitive and whitespace-stripped match")
                return sheet
        
        # If no match found, provide a helpful error message with available options
        available_sheets = ", ".join(workbook.sheetnames)
        raise ValueError(f"Sheet '{target_sheet_name}' not found. Available sheets: {available_sheets}")

    try:
        # Load the workbook with error handling
        try:
            workbook = openpyxl.load_workbook(excel_path)
        except Exception as e:
            raise Exception(f"Failed to load Excel file '{excel_path}': {str(e)}")

        # Find the best matching sheet name
        actual_sheet_name = find_matching_sheet(workbook, sheet_name)
        sheet = workbook[actual_sheet_name]
        
        # Get dimensions and initialize feature array
        max_row = sheet.max_row
        max_col = sheet.max_column
        features = np.zeros((max_row, max_col, 20))
        
        # Process each cell with careful error handling
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                try:
                    cell = sheet.cell(row, col)
                    cell_features = np.zeros(20)
                    
                    # Get string representation of cell value
                    str_value = str(cell.value) if cell.value is not None else ""
                    
                    # 1. Value String Features (indices 0-5)
                    cell_features[0] = 1 if cell.value is not None else 0  # Is non-empty
                    cell_features[1] = len(str_value)  # String length
                    cell_features[2] = sum(c.isdigit() for c in str_value) / max(len(str_value), 1)  # Digit proportion
                    cell_features[3] = sum(c.isalpha() for c in str_value) / max(len(str_value), 1)  # Letter proportion
                    cell_features[4] = 1 if '%' in str_value else 0  # Has percent
                    cell_features[5] = 1 if '.' in str_value else 0  # Has decimal
                    
                    # 2. Data Format Features (indices 6-9)
                    cell_features[6] = 1 if cell.data_type == 'n' else 0  # Is number
                    cell_features[7] = 1 if cell.data_type == 'd' else 0  # Is date
                    cell_features[8] = 1 if isinstance(cell.value, (datetime.time, datetime.datetime)) else 0  # Is time
                    cell_features[9] = len(str(cell.number_format)) if hasattr(cell, 'number_format') and cell.number_format else 0
                    
                    # 3. Cell Format Features (indices 10-17)
                    # Background color
                    if hasattr(cell, 'fill') and cell.fill:
                        cell_features[10] = 1 if (hasattr(cell.fill, 'start_color') and 
                                                cell.fill.start_color and 
                                                cell.fill.start_color.index != '00000000') else 0
                    
                    # Font properties
                    if hasattr(cell, 'font') and cell.font:
                        cell_features[11] = 1 if getattr(cell.font, 'color', None) else 0  # Has font color
                        cell_features[12] = 1 if getattr(cell.font, 'bold', False) else 0  # Is bold
                    
                    # Border properties
                    if hasattr(cell, 'border') and cell.border:
                        cell_features[13] = 1 if getattr(cell.border.left, 'style', None) not in [None, 'none'] else 0
                        cell_features[14] = 1 if getattr(cell.border.right, 'style', None) not in [None, 'none'] else 0
                        cell_features[15] = 1 if getattr(cell.border.top, 'style', None) not in [None, 'none'] else 0
                        cell_features[16] = 1 if getattr(cell.border.bottom, 'style', None) not in [None, 'none'] else 0
                    
                    # Merged cells
                    if hasattr(sheet, 'merged_cells'):
                        cell_features[17] = 1 if any(cell.coordinate in rng for rng in sheet.merged_cells.ranges) else 0
                    
                    # 4. Formula Features (indices 18-19)
                    cell_features[18] = 1 if cell.data_type == 'f' else 0  # Has formula
                    cell_features[19] = 0  # Reserved for future use
                    
                    # Store the features for this cell
                    features[row-1, col-1] = cell_features
                    
                except Exception as e:
                    print(f"Warning: Error processing cell at row {row}, col {col}: {str(e)}")
                    continue
        
        return features
        
    except Exception as e:
        print(f"Error processing file {excel_path}, sheet {sheet_name}: {str(e)}")
        # Return minimal feature array as fallback to allow training to continue
        return np.zeros((1, 1, 20))